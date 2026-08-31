import os
import openpyxl
import json

target = r'C:\DEV\checkOn\창원교회_출석부_20260823.종합시트포함.xlsx'
wb = openpyxl.load_workbook(target, data_only=True)
ws = wb['data_주일']

print('Rows count:', ws.max_row)

address_rows = []
newcomer_rows = []
other_rows = []

for r in range(11, ws.max_row + 1):
    c1 = ws.cell(r, 1).value
    c2 = str(ws.cell(r, 2).value).strip() if ws.cell(r, 2).value is not None else ''
    c3 = str(ws.cell(r, 3).value).strip() if ws.cell(r, 3).value is not None else ''
    c4 = str(ws.cell(r, 4).value).strip() if ws.cell(r, 4).value is not None else ''
    
    if not c4:
        continue
        
    if c1 == '주소록':
        address_rows.append((r, c2, c3, c4))
    elif c1 == '새참자':
        newcomer_rows.append((r, c2, c3, c4))
    else:
        other_rows.append((r, c1, c2, c3, c4))

print(f'주소록: {len(address_rows)}명')
print(f'새참자: {len(newcomer_rows)}명')
print(f'기타: {len(other_rows)}명')

print('\n새참자 전체 목록 in 0823 file:')
for nr in newcomer_rows:
    print(f'  Row {nr[0]}: 구역 {nr[1]}, 인도 {nr[2]}, 이름 {nr[3]}')

if other_rows:
    print('\n기타 행:')
    for o in other_rows:
        print(o)
