import os
import openpyxl

for f in os.listdir('attCheck_ref/founds'):
    if '3분기' in f or '3б' in f:
        path = os.path.join('attCheck_ref/founds', f)
        print('=== File:', f, '===')
        wb = openpyxl.load_workbook(path, data_only=True)
        for s in wb.sheetnames:
            ws = wb[s]
            print(f'\n--- Sheet {s} (rows={ws.max_row}, cols={ws.max_column}) ---')
            r1 = [ws.cell(1, c).value for c in range(1, ws.max_column+1)]
            r2 = [ws.cell(2, c).value for c in range(1, ws.max_column+1)]
            print('  Row 1 (Months):', [(c, v) for c, v in enumerate(r1, 1) if v is not None])
            print('  Row 2 (Types):', [(c, v) for c, v in enumerate(r2, 1) if v is not None])
            # Check how many 'O' or attendance marks exist in 6월, 7월, 8월 columns
            for r in range(3, ws.max_row+1):
                name = ws.cell(r, 2).value
                if name:
                    marks = [(c, ws.cell(r, c).value) for c in range(3, ws.max_column+1) if ws.cell(r, c).value is not None]
                    print(f'    Row {r} ({name}): marks count = {len(marks)}')
            break
