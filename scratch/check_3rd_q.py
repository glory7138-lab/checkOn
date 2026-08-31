import os
import openpyxl

for f in os.listdir('attCheck_ref/founds'):
    if '3' in f or '3분기' in f or '26' in f:
        path = os.path.join('attCheck_ref/founds', f)
        print('FILE:', f)
        wb = openpyxl.load_workbook(path, data_only=True)
        for s in wb.sheetnames:
            ws = wb[s]
            print(f' Sheet {s} (rows={ws.max_row}, cols={ws.max_column})')
            r1 = [ws.cell(1, c).value for c in range(1, ws.max_column+1)]
            r2 = [ws.cell(2, c).value for c in range(1, ws.max_column+1)]
            print('   Row 1:', [v for v in r1 if v is not None])
            print('   Row 2:', [v for v in r2 if v is not None])
            for r in range(3, min(10, ws.max_row+1)):
                row_vals = [ws.cell(r, c).value for c in range(1, ws.max_column+1)]
                print(f'   Row {r}:', [v for v in row_vals if v is not None])
