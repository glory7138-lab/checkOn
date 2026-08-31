import os
import openpyxl

files_to_check = []
for root, dirs, files in os.walk(r'C:\DEV'):
    for f in files:
        if f.endswith('.xlsx') and 'att' in root.lower() or 'founds' in root.lower():
            files_to_check.append(os.path.join(root, f))

for f in files_to_check:
    print('=== File:', f, '===')
    try:
        wb = openpyxl.load_workbook(f, data_only=True)
        for s in wb.sheetnames:
            ws = wb[s]
            print(f'  Sheet: {s}, rows={ws.max_row}, cols={ws.max_column}')
            # check non-empty cells in 7월, 8월 or col headers
            for r in range(1, min(15, ws.max_row+1)):
                vals = [ws.cell(r, c).value for c in range(1, min(30, ws.max_column+1))]
                if any(vals):
                    print(f'    Row {r}: {[v for v in vals if v is not None][:10]}')
    except Exception as e:
        print('  Error:', e)
