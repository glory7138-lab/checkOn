import openpyxl

target = r'C:\DEV\checkOn\창원교회_출석부_20260823.종합시트포함.xlsx'
wb = openpyxl.load_workbook(target, data_only=True)

ws = wb['data_주일']

print('--- data_주일 Row 308 ---')
for c in range(1, 5):
    val = ws.cell(308, c).value
    codes = [hex(ord(ch)) for ch in str(val)]
    print(f'Col {c}: {repr(val)} -> {codes}')

print('\n--- data_주일 Row 309 ---')
for c in range(1, 5):
    val = ws.cell(309, c).value
    codes = [hex(ord(ch)) for ch in str(val)]
    print(f'Col {c}: {repr(val)} -> {codes}')

if '출석부 추가표시 명단' in wb.sheetnames:
    ws_g = wb['출석부 추가표시 명단']
    print('\n--- 출석부 추가표시 명단 Row 1 ---')
    for c in range(1, 6):
        val = ws_g.cell(1, c).value
        codes = [hex(ord(ch)) for ch in str(val)] if val else []
        print(f'Col {c}: {repr(val)} -> {codes}')
    print('\n--- 출석부 추가표시 명단 Row 2 ---')
    for c in range(1, 6):
        val = ws_g.cell(2, c).value
        codes = [hex(ord(ch)) for ch in str(val)] if val else []
        print(f'Col {c}: {repr(val)} -> {codes}')
