import openpyxl

target = r'C:\DEV\checkOn\창원교회_출석부_20260823.종합시트포함.xlsx'
wb = openpyxl.load_workbook(target, data_only=True)

ws = wb['data_주일']

print('--- data_주일 Row 308 ---')
for c in range(1, 5):
    print(f'Col {c}: {ws.cell(308, c).value}')

print('\n--- data_주일 Rows 309 to 315 ---')
for r in range(309, 316):
    print(f'Row {r}: Col1={ws.cell(r, 1).value}, Col2={ws.cell(r, 2).value}, Col3={ws.cell(r, 3).value}, Col4={ws.cell(r, 4).value}')

if '출석부 추가표시 명단' in wb.sheetnames:
    ws_g = wb['출석부 추가표시 명단']
    print('\n--- 출석부 추가표시 명단 Row 1 ---')
    for c in range(1, 6):
        print(f'Col {c}: {ws_g.cell(1, c).value}')
    print('--- 출석부 추가표시 명단 Rows 2 to 6 ---')
    for r in range(2, 7):
        print(f'Row {r}: Col3={ws_g.cell(r, 3).value}, Col4={ws_g.cell(r, 4).value}, Col5={ws_g.cell(r, 5).value}')
