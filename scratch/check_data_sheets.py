import os
import openpyxl

for f in os.listdir('attCheck_ref/founds'):
    if '20260607' in f:
        path = os.path.join('attCheck_ref/founds', f)
        print('=== File:', f, '===')
        wb = openpyxl.load_workbook(path, data_only=True)
        for s in wb.sheetnames:
            ws = wb[s]
            print(f'\n--- Sheet {s} (max_row={ws.max_row}, max_col={ws.max_column}) ---')
            # Check non-empty cells
            filled_cells = 0
            for r in range(1, ws.max_row+1):
                for c in range(1, ws.max_column+1):
                    if ws.cell(r, c).value is not None:
                        filled_cells += 1
            print(f'  Filled cells: {filled_cells}')
            
            # For data_주일 specifically, let's check cols 32 to 57
            if 'data' in s:
                for c in range(30, min(45, ws.max_column+1)):
                    header_date = ws.cell(8, c).value
                    col_vals = [ws.cell(r, c).value for r in range(11, ws.max_row+1) if ws.cell(r, c).value is not None]
                    print(f'  Col {c} ({header_date}): {len(col_vals)} non-empty cells -> {col_vals[:5]}')
