import openpyxl

target = r'C:\DEV\checkOn\창원교회_출석부_20260823.종합시트포함.xlsx'
wb = openpyxl.load_workbook(target, data_only=False)
ws = wb['data_주일']

print('Formula in Row 9 Col 5:', ws.cell(9, 5).value)
print('Formula in Row 9 Col 31:', ws.cell(9, 31).value)
print('Formula in Row 9 Col 42:', ws.cell(9, 42).value)
