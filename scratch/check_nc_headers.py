import openpyxl

target = r'C:\DEV\checkOn\창원교회_출석부_20260823.종합시트포함.xlsx'
wb = openpyxl.load_workbook(target, data_only=True)

ws = wb['data_주일']

print('--- data_주일 Row 10 headers ---')
r10 = [ws.cell(10, c).value for c in range(1, 10)]
print('Row 10 (1..9):', r10)

print('\n--- data_주일 Row 308 (Newcomer section header) ---')
r308 = [ws.cell(308, c).value for c in range(1, 10)]
print('Row 308 (1..9):', r308)

print('\n--- data_주일 Rows 309 to 320 ---')
for r in range(309, 321):
    c1 = ws.cell(r, 1).value
    c2 = ws.cell(r, 2).value
    c3 = ws.cell(r, 3).value
    c4 = ws.cell(r, 4).value
    print(f'Row {r}: Col1={c1}, Col2={c2}, Col3={c3}, Col4={c4}')

if '출석부 추가표시 명단' in wb.sheetnames:
    ws_guide = wb['출석부 추가표시 명단']
    print('\n--- 출석부 추가표시 명단 (Rows 1 to 10) ---')
    for r in range(1, 11):
        vals = [ws_guide.cell(r, c).value for c in range(1, 6)]
        print(f'Row {r}: {vals}')
