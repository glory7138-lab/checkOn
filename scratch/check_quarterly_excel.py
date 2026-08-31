import os
import openpyxl

for fname in [r'attCheck_ref\founds\263бӴȸ⼮.xlsx', r'attCheck_ref\founds\2026_Ӵȸ⼮üũ.xlsx']:
    for f in os.listdir('attCheck_ref/founds'):
        if f.endswith('.xlsx') and ('26' in f or '2026' in f) and '20260607' not in f:
            full_path = os.path.join('attCheck_ref/founds', f)
            print('========================================')
            print('Inspecting:', f)
            wb = openpyxl.load_workbook(full_path, data_only=True)
            for s in wb.sheetnames:
                ws = wb[s]
                print(f'--- Sheet {s} (rows: {ws.max_row}, cols: {ws.max_column}) ---')
                r1 = [ws.cell(1, c).value for c in range(1, ws.max_column+1)]
                r2 = [ws.cell(2, c).value for c in range(1, ws.max_column+1)]
                print('  Row 1:', [v for v in r1 if v is not None])
                print('  Row 2:', [v for v in r2 if v is not None][:15])
                # Show sample member rows
                for r in range(3, min(8, ws.max_row+1)):
                    row_vals = [ws.cell(r, c).value for c in range(1, ws.max_column+1)]
                    print(f'  Row {r}:', [v for v in row_vals if v is not None])
            break
