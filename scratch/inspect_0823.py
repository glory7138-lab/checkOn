import os
import openpyxl

target = r'C:\DEV\checkOn\창원교회_출석부_20260823.종합시트포함.xlsx'
wb = openpyxl.load_workbook(target, data_only=True)
print('Sheet names:', wb.sheetnames)

if 'data_주일' in wb.sheetnames:
    ws = wb['data_주일']
    print(f'data_주일 dimensions: max_row={ws.max_row}, max_column={ws.max_column}')
    
    # Check date headers (row 7, 8, 9, 10)
    print('\nDate columns:')
    for c in range(5, ws.max_column + 1):
        r8 = ws.cell(8, c).value
        r9 = ws.cell(9, c).value
        r10 = ws.cell(10, c).value
        print(f'  Col {c}: r8={r8}, r9(총출석)={r9}, r10={r10}')
        
    print('\nChecking row types (Col 1):')
    row_types = {}
    for r in range(11, ws.max_row + 1):
        c1 = ws.cell(r, 1).value
        row_types[c1] = row_types.get(c1, 0) + 1
    print('Row types count:', row_types)
