import os
import openpyxl
import json
import datetime

target = r'C:\DEV\checkOn\창원교회_출석부_20260823.종합시트포함.xlsx'
wb = openpyxl.load_workbook(target, data_only=True)
ws = wb['data_주일']

def get_sunday_date(col):
    if col == 5: return '2025-12-07'
    elif col == 6: return '2025-12-14'
    elif col == 7: return '2025-12-21'
    elif col == 8: return '2025-12-28'
    else:
        base_2026 = datetime.date(2026, 1, 4)
        return (base_2026 + datetime.timedelta(weeks=(col - 9))).strftime('%Y-%m-%d')

all_data = []
for r in range(11, ws.max_row + 1):
    c1 = ws.cell(r, 1).value
    c2 = str(ws.cell(r, 2).value).strip() if ws.cell(r, 2).value is not None else ''
    c3 = str(ws.cell(r, 3).value).strip() if ws.cell(r, 3).value is not None else ''
    c4 = str(ws.cell(r, 4).value).strip() if ws.cell(r, 4).value is not None else ''
    
    if not c4 or r == 308 or c1 == '구분':
        continue
        
    row_type = 'address' if c1 == '주소록' else 'newcomer'
    
    att_dates = []
    for c in range(5, 43): # Col 5 (2025-12-07) to Col 42 (2026-08-23)
        val = ws.cell(r, c).value
        if val is not None and str(val).strip() in ['1', '1.0', 'O', 'o', 'v', 'V']:
            att_dates.append(get_sunday_date(col=c))
            
    all_data.append({
        'row': r,
        'type': row_type,
        'area_code': c2,
        'guide_name': c3,
        'name': c4,
        'attended_dates': att_dates
    })

print(f'Total parsed entries: {len(all_data)}')
print(f'Addressbook entries: {len([x for x in all_data if x["type"] == "address"])}')
print(f'Newcomer entries: {len([x for x in all_data if x["type"] == "newcomer"])}')

# Date-wise summary comparison with Row 9
date_summary = {}
for item in all_data:
    for d in item['attended_dates']:
        date_summary[d] = date_summary.get(d, 0) + 1

print('\nDate-wise attendance count check:')
mismatches = 0
for c in range(5, 43):
    d_str = get_sunday_date(c)
    expected_sum = int(ws.cell(9, c).value or 0)
    actual_sum = date_summary.get(d_str, 0)
    match_str = 'OK' if expected_sum == actual_sum else f'MISMATCH (Expected: {expected_sum}, Actual: {actual_sum})'
    if expected_sum != actual_sum:
        mismatches += 1
    print(f'  Col {c:2d} ({d_str}): {actual_sum:>3d}명  [{match_str}]')

print(f'\nTotal mismatches: {mismatches}')

with open('scratch_parsed_0823.json', 'w', encoding='utf-8') as f:
    json.dump(all_data, f, ensure_ascii=False, indent=2)
