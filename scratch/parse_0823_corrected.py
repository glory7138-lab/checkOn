import os
import openpyxl
import json
import datetime

target = r'C:\DEV\checkOn\창원교회_출석부_20260823.종합시트포함.xlsx'
wb = openpyxl.load_workbook(target, data_only=True)
ws = wb['data_주일']

def get_sunday_date(col):
    if col == 5: return '2025-12-07'
    elif col == 6: return '2025-12-14'
    elif col == 7: return '2025-12-21'
    elif col == 8: return '2025-12-28'
    else:
        base_2026 = datetime.date(2026, 1, 4)
        return (base_2026 + datetime.timedelta(weeks=(col - 9))).strftime('%Y-%m-%d')

all_data = []
for r in range(11, ws.max_row + 1):
    c1 = ws.cell(r, 1).value
    c2 = str(ws.cell(r, 2).value).strip() if ws.cell(r, 2).value is not None else ''
    c3 = str(ws.cell(r, 3).value).strip() if ws.cell(r, 3).value is not None else ''
    c4 = str(ws.cell(r, 4).value).strip() if ws.cell(r, 4).value is not None else ''
    
    if r == 308 or c1 == '구분':
        continue
        
    if c1 == '주소록':
        if not c4: continue
        row_type = 'address'
        member_name = c4
        guide_name = ''
    elif c1 == '새참자':
        if not c3 and not c4: continue
        row_type = 'newcomer'
        member_name = c3  # Col 3 is Newcomer Name (성명)
        guide_name = c4   # Col 4 is Guide Name (인도자)
    else:
        continue
    
    att_dates = []
    for c in range(5, 43): # Col 5 (2025-12-07) to Col 42 (2026-08-23)
        val = ws.cell(r, c).value
        if val is not None and str(val).strip() in ['1', '1.0', 'O', 'o', 'v', 'V']:
            att_dates.append(get_sunday_date(col=c))
            
    all_data.append({
        'row': r,
        'type': row_type,
        'area_code': c2,
        'guide_name': guide_name,
        'name': member_name,
        'attended_dates': att_dates
    })

print(f'Total parsed entries: {len(all_data)}')
print(f'Addressbook entries: {len([x for x in all_data if x["type"] == "address"])}')
print(f'Newcomer entries: {len([x for x in all_data if x["type"] == "newcomer"])}')

print('\nCorrected Newcomers list (Sample first 15):')
for nc in [x for x in all_data if x['type'] == 'newcomer'][:15]:
    print(f'  Row {nc["row"]}: [구역 {nc["area_code"]}] 새참자: {nc["name"]}, 인도자: {nc["guide_name"]} (출석 {len(nc["attended_dates"])}회)')

with open('scratch_parsed_0823_corrected.json', 'w', encoding='utf-8') as f:
    json.dump(all_data, f, ensure_ascii=False, indent=2)
