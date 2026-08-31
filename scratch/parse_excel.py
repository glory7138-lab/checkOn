import os
import openpyxl
import json
import datetime

wb = None
for f in os.listdir('attCheck_ref/founds'):
    if f.endswith('.xlsx') and '20260607' in f:
        wb = openpyxl.load_workbook(os.path.join('attCheck_ref/founds', f), data_only=True)
        break

ws = wb['data_\uc8fc\uc77c']

def get_sunday_date(col):
    if col == 5: return '2025-12-07'
    elif col == 6: return '2025-12-14'
    elif col == 7: return '2025-12-21'
    elif col == 8: return '2025-12-28'
    else:
        base_2026 = datetime.date(2026, 1, 4)
        return (base_2026 + datetime.timedelta(weeks=(col - 9))).strftime('%Y-%m-%d')

all_data = []
for r in range(11, 326):
    c1 = ws.cell(r, 1).value
    c2 = str(ws.cell(r, 2).value).strip() if ws.cell(r, 2).value is not None else ''
    c3 = str(ws.cell(r, 3).value).strip() if ws.cell(r, 3).value is not None else ''
    c4 = str(ws.cell(r, 4).value).strip() if ws.cell(r, 4).value is not None else ''
    
    if not c4 or r == 309:
        continue
        
    row_type = 'address' if c1 == '\uc8fc\uc18c\ub85d' else 'newcomer'
    
    att_dates = []
    for c in range(5, 43): # up to 2026-08-23
        val = ws.cell(r, c).value
        if val is not None and str(val).strip() in ['1', '1.0', 'O', 'o', 'v', 'V']:
            att_dates.append(get_sunday_date(col=c))
            
    all_data.append({
        'row': r,
        'type': row_type,
        'area_code': c2,
        'guide_name': c3,
        'name': c4,
        'attended_dates': att_dates
    })

print(f'Total parsed entries: {len(all_data)}')
print(f'Addressbook entries: {len([x for x in all_data if x["type"] == "address"])}')
print(f'Newcomer entries: {len([x for x in all_data if x["type"] == "newcomer"])}')

# Date-wise summary
date_summary = {}
for item in all_data:
    for d in item['attended_dates']:
        date_summary[d] = date_summary.get(d, 0) + 1

print('\nDate-wise attendance counts from Excel:')
for d in sorted(date_summary.keys()):
    print(f'  {d}: {date_summary[d]}')

with open('scratch_parsed_attendance.json', 'w', encoding='utf-8') as f:
    json.dump(all_data, f, ensure_ascii=False, indent=2)
